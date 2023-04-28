from typing import List
from datetime import date
from doublons import verifier_doublons
from scraping import extraire_articles_recherche
from google_search import google_search
from traduction import traduire_articles
from summary import resumer_articles
import openpyxl
import xlwt



def rechercher_articles(mots_cles: str, date_debut: date, date_fin: date, extraire_articles: bool, generer_fichier: bool, verifier_doublons: bool, resumer_articles: bool, traduire_articles: bool, cx: str, nb_resultats: int = 10, langue: str = "fr") -> List[dict]:
    """
    Effectue une recherche d'articles sur Google News pour les mots clés et la plage de dates spécifiés.
    """
    if extraire_articles:
        articles = extraire_articles_recherche(mots_cles, date_debut, date_fin)
    else:
        articles = google_search(mots_cles, cx, nb_resultats)
    if verifier_doublons:
        articles = verifier_doublons(articles)
    if resumer_articles:
        articles = resumer_articles(articles)
    if traduire_articles:
        articles = traduire_articles(articles, langue)
    if generer_fichier:
        generer_fichier_excel(articles)
    return articles

def generer_fichier_excel(articles: List[dict]):
    """
    Génère un fichier Excel contenant les informations des articles.
    """
    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # Ajout des en-têtes
    ws["A1"] = "Titre"
    ws["B1"] = "URL"
    ws["C1"] = "Date"
    ws["D1"] = "Source"
    ws["E1"] = "Résumé"

    # Ajout des données des articles
    row = 2
    for article in articles:
        ws.cell(row=row, column=1, value=article["titre"])
        ws.cell(row=row, column=2, value=article["url"])
        ws.cell(row=row, column=3, value=article["date"])
        ws.cell(row=row, column=4, value=article["source"])
        ws.cell(row=row, column=5, value=article["resume"])
        row += 1

    # Enregistrement du fichier Excel
    wb.save("articles.xlsx")

